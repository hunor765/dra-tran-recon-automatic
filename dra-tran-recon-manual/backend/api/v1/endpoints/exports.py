"""Data export endpoints for job results."""
import csv
import io
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException, status, Request
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from core.auth import get_current_user, require_admin
from core.database import get_db
from core.rate_limiter import limiter, RateLimits
from models.job import Job as JobModel
from models.client import Client as ClientModel

router = APIRouter()


@router.get("/jobs/{job_id}/export")
@limiter.limit(RateLimits.GET)
async def export_job_results(
    request: Request,
    job_id: int,
    format: Literal["csv", "json"] = "csv",
    data_type: Literal["missing", "all"] = "missing",
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user)
):
    """Export job results as CSV or JSON.
    
    Args:
        job_id: The job ID to export
        format: Export format (csv or json)
        data_type: Type of data to export (missing transactions only, or all data)
    """
    # Get job with client info
    result = await db.execute(
        select(JobModel, ClientModel.name.label("client_name"))
        .join(ClientModel, JobModel.client_id == ClientModel.id)
        .where(JobModel.id == job_id)
    )
    row = result.first()
    
    if not row:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job, client_name = row
    
    # Check job is completed
    if job.status != "completed":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot export job with status: {job.status}. Job must be completed."
        )
    
    if not job.result_summary:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Job has no result data to export"
        )
    
    # Generate filename
    safe_client_name = "".join(c for c in (client_name or "client") if c.isalnum() or c in "_-").lower()
    timestamp = job.completed_at.strftime("%Y%m%d_%H%M") if job.completed_at else "unknown"
    filename = f"{safe_client_name}_missing_transactions_{timestamp}"
    
    missing_ids = job.result_summary.get("missing_ids", [])
    
    if format == "json":
        import json
        
        export_data = {
            "job_id": job.id,
            "client_name": client_name,
            "exported_at": datetime.utcnow().isoformat(),
            "date_range": job.result_summary.get("date_range"),
            "summary": {
                "match_rate": job.result_summary.get("match_rate"),
                "total_backend_records": job.result_summary.get("backend_records"),
                "total_ga4_records": job.result_summary.get("ga4_records"),
                "missing_count": len(missing_ids),
            },
            "missing_transaction_ids": missing_ids,
        }
        
        return StreamingResponse(
            io.BytesIO(json.dumps(export_data, indent=2).encode()),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={filename}.json"}
        )
    
    else:  # CSV format
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            "Transaction ID",
            "Job ID",
            "Client Name",
            "Date Range Start",
            "Date Range End",
            "Match Rate (%)",
            "Exported At"
        ])
        
        # Write data rows
        date_range = job.result_summary.get("date_range", {})
        match_rate = job.result_summary.get("match_rate", 0)
        from datetime import datetime
        
        if missing_ids:
            for i, tx_id in enumerate(missing_ids):
                writer.writerow([
                    tx_id,
                    job.id,
                    client_name,
                    date_range.get("start_date", ""),
                    date_range.get("end_date", ""),
                    match_rate,
                    datetime.utcnow().isoformat() if i == 0 else ""  # Only show once
                ])
        else:
            # Write empty row with metadata
            writer.writerow([
                "N/A - No missing transactions",
                job.id,
                client_name,
                date_range.get("start_date", ""),
                date_range.get("end_date", ""),
                match_rate,
                datetime.utcnow().isoformat()
            ])
        
        output.seek(0)
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode()),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={filename}.csv",
                "Content-Type": "text/csv; charset=utf-8"
            }
        )


@router.get("/jobs/{job_id}/export/excel")
@limiter.limit(RateLimits.GET)
async def export_job_excel(
    request: Request,
    job_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user)
):
    """Export job results as Excel file.
    
    Note: This requires openpyxl to be installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Excel export is not available. Please install openpyxl: pip install openpyxl"
        )
    
    # Get job with client info
    result = await db.execute(
        select(JobModel, ClientModel.name.label("client_name"))
        .join(ClientModel, JobModel.client_id == ClientModel.id)
        .where(JobModel.id == job_id)
    )
    row = result.first()
    
    if not row:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job, client_name = row
    
    if job.status != "completed":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot export job with status: {job.status}"
        )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Missing Transactions"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="DD3333", end_color="DD3333", fill_type="solid")
    
    # Summary section
    ws["A1"] = "Reconciliation Report"
    ws["A1"].font = Font(bold=True, size=16)
    
    summary_data = [
        ["Client:", client_name],
        ["Job ID:", job.id],
        ["Date Range:", f"{job.result_summary.get('date_range', {}).get('start_date', 'N/A')} to {job.result_summary.get('date_range', {}).get('end_date', 'N/A')}"],
        ["Match Rate:", f"{job.result_summary.get('match_rate', 0)}%"],
        ["Total Backend Records:", job.result_summary.get("backend_records", 0)],
        ["Total GA4 Records:", job.result_summary.get("ga4_records", 0)],
        ["Missing Count:", len(job.result_summary.get("missing_ids", []))],
    ]
    
    for i, (label, value) in enumerate(summary_data, start=3):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = value
    
    # Missing transactions section
    start_row = len(summary_data) + 5
    
    ws[f"A{start_row}"] = "Missing Transaction IDs"
    ws[f"A{start_row}"].font = Font(bold=True, size=14)
    
    # Headers
    header_row = start_row + 1
    ws[f"A{header_row}"] = "#"
    ws[f"B{header_row}"] = "Transaction ID"
    
    for cell in [f"A{header_row}", f"B{header_row}"]:
        ws[cell].font = header_font
        ws[cell].fill = header_fill
        ws[cell].alignment = Alignment(horizontal="center")
    
    # Data
    missing_ids = job.result_summary.get("missing_ids", [])
    for i, tx_id in enumerate(missing_ids, start=1):
        row_num = header_row + i
        ws[f"A{row_num}"] = i
        ws[f"B{row_num}"] = tx_id
    
    # Adjust column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40
    
    # Save to buffer
    from datetime import datetime
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    safe_client_name = "".join(c for c in (client_name or "client") if c.isalnum() or c in "_-").lower()
    timestamp = job.completed_at.strftime("%Y%m%d_%H%M") if job.completed_at else "unknown"
    filename = f"{safe_client_name}_reconciliation_{timestamp}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
